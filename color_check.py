from openpyxl import load_workbook

# --- [사용자 설정] ---
# 1. 색상을 확인할 엑셀 파일의 전체 경로를 입력하세요.
FILE_PATH = r"C:\Users\user\Desktop\전기테스트.xlsx" 

# 2. 해당 셀이 있는 시트의 정확한 이름을 입력하세요.
SHEET_NAME = "서울"

# 3. 색상을 확인할 셀의 주소를 입력하세요 (예: "H11").
CELL_ADDRESS = "L156"
# --------------------

try:
    # 서식을 포함하여 엑셀 파일을 엽니다.
    workbook = load_workbook(filename=FILE_PATH, data_only=False)
    
    # 지정된 시트를 선택합니다.
    if SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SHEET_NAME]
    else:
        print(f"[오류] '{SHEET_NAME}' 시트를 찾을 수 없습니다. 시트 이름을 확인해주세요.")
        print(f"  - 사용 가능한 시트 목록: {workbook.sheetnames}")
        exit()

    # 지정된 셀을 선택합니다.
    cell = sheet[CELL_ADDRESS]
    
    # 셀의 채우기(fill) 정보와 색상(fgColor) 객체를 가져옵니다.
    color_obj = cell.fill.fgColor if cell.fill else None

    print(f"\n--- [{SHEET_NAME} 시트의 {CELL_ADDRESS} 셀 색상 정보] ---")
    
    if color_obj:
        print(f" - Type: {color_obj.type}")
        # RGB 값은 16진수 형태(예: FF123456)로 나타납니다.
        print(f" - RGB: {color_obj.rgb}") 
        # 테마 색상일 경우, 테마 번호가 나타납니다.
        print(f" - Theme: {color_obj.theme}")
        # 색상의 밝기/어둡기를 조절하는 값입니다.
        print(f" - Tint: {color_obj.tint}")
    else:
        print(" - 해당 셀에는 색상 정보가 없습니다.")
        
    print("------------------------------------------\n")

except FileNotFoundError:
    print(f"[오류] 파일을 찾을 수 없습니다: {FILE_PATH}")
except Exception as e:
    print(f"[오류] 파일을 처리하는 중 오류가 발생했습니다: {e}")