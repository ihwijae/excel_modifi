"""Utilities for extracting business numbers and related metadata from the Excel DB."""

from __future__ import annotations

import re
from collections import OrderedDict
from dataclasses import dataclass
from typing import Iterable, List, Optional, Sequence

from openpyxl import load_workbook

from config import RELATIVE_OFFSETS

_BIZ_NO_REGEX = re.compile(r"(\d{3})[- ]?(\d{2})[- ]?(\d{5})")
_LABEL_KEYWORD = "사업자번호"
_DATA_START_COLUMN = 2  # 열 A는 라벨, B부터 실제 데이터가 배치되어 있음


@dataclass
class CompanyEntry:
    """단일 업체(열)의 사업자번호와 부가 정보를 담는 구조체."""

    biz_no: str
    company_name: str = ""
    ceo_name: str = ""
    sheet_name: str = ""
    column_index: int = 0
    base_row_index: int = 0


def normalize_biz_no(raw: Optional[str]) -> Optional[str]:
    """Normalize any raw input into `000-00-00000` form if possible."""
    if not raw:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) != 10:
        return None
    return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"


def extract_company_entries_from_excel(
    excel_path: str,
    sheet_names: Optional[Sequence[str]] = None,
) -> List[CompanyEntry]:
    """
    워크북에서 사업자번호/업체명/대표자명 정보를 모두 읽어온다.

    기존 방식처럼 시트 전체를 훑어 숫자를 찾는 것이 아니라,
    '사업자번호' 라벨이 붙은 행을 기준으로 동일한 열에 있는 데이터를 읽어서
    중복 없이 정확한 업체 목록을 만든다.
    """
    workbook = load_workbook(filename=excel_path, data_only=True)
    try:
        target_sheet_names: Iterable[str]
        if sheet_names:
            target_sheet_names = sheet_names
        else:
            target_sheet_names = workbook.sheetnames

        entries: "OrderedDict[str, CompanyEntry]" = OrderedDict()

        for sheet_name in target_sheet_names:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
            sheet = workbook[sheet_name]
            biz_rows = _find_biz_no_rows(sheet)

            # 라벨을 찾지 못하면 이전 방식으로라도 숫자를 추출하도록 폴백
            if not biz_rows:
                for row in sheet.iter_rows(values_only=True):
                    for cell_value in row:
                        normalized = _extract_single_biz_no(cell_value)
                        if normalized and normalized not in entries:
                            entries[normalized] = CompanyEntry(
                                biz_no=normalized,
                                sheet_name=sheet_name,
                            )
                continue

            company_offset = RELATIVE_OFFSETS.get("회사명", -2)
            ceo_offset = RELATIVE_OFFSETS.get("대표자", -1)

            for biz_row in biz_rows:
                for col in range(_DATA_START_COLUMN, sheet.max_column + 1):
                    cell_value = sheet.cell(row=biz_row, column=col).value
                    normalized = _extract_single_biz_no(cell_value)
                    if not normalized:
                        continue

                    company_name = _read_related_cell(sheet, biz_row, col, company_offset)
                    ceo_name = _read_related_cell(sheet, biz_row, col, ceo_offset)

                    entry = entries.get(normalized)
                    if not entry:
                        entries[normalized] = CompanyEntry(
                            biz_no=normalized,
                            company_name=company_name,
                            ceo_name=ceo_name,
                            sheet_name=sheet_name,
                            column_index=col,
                            base_row_index=biz_row,
                        )
                    else:
                        # 이미 있는 경우, 비어 있는 필드만 채워준다.
                        if not entry.company_name and company_name:
                            entry.company_name = company_name
                        if not entry.ceo_name and ceo_name:
                            entry.ceo_name = ceo_name
                        if not entry.sheet_name:
                            entry.sheet_name = sheet_name
                        if not entry.column_index:
                            entry.column_index = col
                        if not entry.base_row_index:
                            entry.base_row_index = biz_row

        return list(entries.values())
    finally:
        workbook.close()


def extract_biz_numbers_from_excel(
    excel_path: str,
    sheet_names: Optional[Sequence[str]] = None,
) -> List[str]:
    """
    기존 호출부와의 호환성을 위해 사업자번호 문자열만 따로 반환하는 래퍼.
    """
    return [entry.biz_no for entry in extract_company_entries_from_excel(excel_path, sheet_names)]


def _find_biz_no_rows(sheet) -> List[int]:
    """시트 내에서 '사업자번호' 라벨이 등장하는 모든 행 번호를 찾는다."""
    rows: List[int] = []
    for row_idx in range(1, sheet.max_row + 1):
        label_value = sheet.cell(row=row_idx, column=1).value
        if not label_value:
            continue
        normalized = _normalize_label(str(label_value))
        if not normalized:
            continue
        if _LABEL_KEYWORD in normalized.replace("등록", ""):
            rows.append(row_idx)
    return rows


def _read_related_cell(sheet, base_row: int, column: int, offset: int) -> str:
    """기준 행에서 상대 위치에 있는 셀 값을 읽어 문자열로 반환."""
    target_row = base_row + offset
    if target_row < 1 or target_row > sheet.max_row:
        return ""
    value = sheet.cell(row=target_row, column=column).value
    if value is None:
        return ""
    return str(value).strip()


def _normalize_label(value: str) -> str:
    return re.sub(r"\s+", "", value).strip()


def _extract_single_biz_no(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        normalized = normalize_biz_no(f"{int(value):010d}")
        return normalized
    text = str(value).strip()
    if not text:
        return None
    normalized = normalize_biz_no(text)
    if normalized:
        return normalized
    match = _BIZ_NO_REGEX.search(text)
    if match:
        joined = "".join(match.groups())
        return normalize_biz_no(joined)
    return None

