from openpyxl import load_workbook
from config import RELATIVE_OFFSETS, COLUMN_MAP, RATIO_THRESHOLDS
from openpyxl.styles import PatternFill, Color, Font
from datetime import datetime
import re


def get_sheet_names(excel_path):
    """
    엑셀 파일의 모든 시트 이름을 리스트로 반환합니다.
    """
    try:
        workbook = load_workbook(filename=excel_path, read_only=True)
        return workbook.sheetnames, None
    except Exception as e:
        return None, f"엑셀 파일 열기 오류: {e}"


def find_company_data(excel_path, biz_no_to_find):
    """
    엑셀 파일에서 업체를 찾아, 값과 함께 셀 배경색 정보도 반환합니다.
    """
    try:
        workbook = load_workbook(filename=excel_path, data_only=False)
    except Exception as e:
        return None, f"엑셀 파일 열기 오류: {e}"

    target_row, target_col, target_sheet_name = -1, -1, None;
    found = False
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None: continue
                if str(cell.value).strip().replace('-', '') == biz_no_to_find.strip().replace('-', ''):
                    target_row, target_col, target_sheet_name = cell.row, cell.column, sheet_name
                    found = True;
                    break
            if found: break
        if found: break

    if not found:
        return None, None

    sheet = workbook[target_sheet_name]
    found_data = {}
    for key, excel_label in COLUMN_MAP.items():
        if excel_label in RELATIVE_OFFSETS:
            row_offset = RELATIVE_OFFSETS[excel_label]
            read_row = target_row + row_offset
            if 1 <= read_row <= sheet.max_row and 1 <= target_col <= sheet.max_column:
                cell = sheet.cell(row=read_row, column=target_col)
                value = cell.value
                color_hex = "#FFFFFF"
                if cell.fill and cell.fill.fgColor:
                    color_info = cell.fill.fgColor
                    if color_info.type == 'theme':
                        if color_info.theme == 6: color_hex = "#E2EFDA"
                        elif color_info.theme == 3: color_hex = "#DDEBF7"
                    elif color_info.type == 'rgb' and isinstance(color_info.rgb, str):
                        hex_val = color_info.rgb
                        color_hex = f"#{hex_val[2:]}" if len(hex_val) == 8 and hex_val.startswith("FF") else f"#{hex_val}"
                if color_hex == '#00000000': color_hex = '#FFFFFF'
                found_data[key] = {'value': value, 'color': color_hex}
            else:
                found_data[key] = {'value': 'N/A', 'color': '#FFFFFF'}

    # [핵심] 찾은 데이터에 '지역' 정보(시트 이름)를 추가
    found_data['지역'] = {'value': target_sheet_name, 'color': '#FFFFFF'}
    return found_data, None


def update_company_data(excel_path, biz_no_to_find, update_data, db_type):
    """
    엑셀에서 업체를 찾아 데이터를 업데이트하고, 조건에 따라 서식을 변경합니다.
    """
    try:
        workbook = load_workbook(filename=excel_path)
    except Exception as e:
        return None, f"엑셀 파일 열기 오류: {e}"

    THEME_GREEN_COLOR = Color(type='theme', theme=6, tint=0.7999816888943144)
    GREEN_FILL = PatternFill(fgColor=THEME_GREEN_COLOR, fill_type="solid")
    DEFAULT_FONT = Font(color="000000", bold=False, size=9)
    HIGHLIGHT_FONT = Font(color="FF0000", bold=True, size=9)

    target_row, target_col, target_sheet_name = -1, -1, None;
    found = False
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None: continue
                if str(cell.value).strip().replace('-', '') == biz_no_to_find.strip().replace('-', ''):
                    target_row, target_col, target_sheet_name = cell.row, cell.column, sheet_name
                    found = True;
                    break
            if found: break
        if found: break
    if not found: return None, f"엑셀 파일에서 사업자번호 '{biz_no_to_find}'를 찾을 수 없습니다."

    sheet = workbook[target_sheet_name]
    updated_log = []
    for key, excel_label in COLUMN_MAP.items():
        if excel_label in RELATIVE_OFFSETS:
            row_offset = RELATIVE_OFFSETS[excel_label]
            update_row = target_row + row_offset
            if 1 <= update_row <= sheet.max_row and 1 <= target_col <= sheet.max_column:
                cell = sheet.cell(row=update_row, column=target_col)
                if key not in ['상호', '신용평가']: cell.fill = GREEN_FILL
                if key in update_data and update_data[key]:
                    # 업체명(상호)은 기존 셀 폰트를 보존한다.
                    if key != '?곹샇':
                        cell.font = DEFAULT_FONT
                    value_str = str(update_data[key]).replace(",", "").replace("%", "")
                    try:
                        numeric_value = 0
                        if '비율' in key:
                            numeric_value = float(value_str)
                            cell.value = numeric_value / 100.0
                            cell.number_format = '0.00%'
                        elif key in ['시평액', '3년실적', '5년실적']:
                            cell.value = int(float(value_str)) * 1000
                        else:
                            cell.value = update_data[key]
                        updated_log.append(excel_label)
                        if db_type and key in ['부채비율', '유동비율']:
                            thresholds = RATIO_THRESHOLDS.get(db_type, {}).get(key, {})
                            if 'max' in thresholds and numeric_value > thresholds['max']: cell.font = HIGHLIGHT_FONT
                            elif 'min' in thresholds and numeric_value < thresholds['min']: cell.font = HIGHLIGHT_FONT
                    except (ValueError, TypeError):
                        pass
    try:
        workbook.save(excel_path)
        return updated_log, None
    except Exception as e:
        return None, f"엑셀 파일 저장 오류: {e}"


def add_new_company_data(excel_path, new_data, company_name, sheet_name, db_type):
    """
    [수정] 신규 업체 데이터를 추가할 때, '회사명'은 회색/굵게, 나머지는 초록색으로 서식을 적용합니다.
    """
    try:
        workbook = load_workbook(filename=excel_path)
        if sheet_name not in workbook.sheetnames:
            return None, f"엑셀 파일에 '{sheet_name}' 시트가 존재하지 않습니다."
        sheet = workbook[sheet_name]
    except Exception as e:
        return None, f"엑셀 파일 열기 오류: {e}"

    # --- 서식 정의 ---
    GREEN_FILL = PatternFill(fgColor=Color(type='theme', theme=6, tint=0.7999816888943144), fill_type="solid")
    GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    BOLD_FONT = Font(bold=True, size=12)
    DEFAULT_FONT = Font(color="000000", bold=False, size=9)

    # 명세서 기반 상수 정의
    START_COL, END_COL = 2, 13
    COMPANY_LABEL, REMARKS_LABEL = "회사명", "비고"

    # 1. 구간 탐지
    sections = []
    start_row = -1
    for row in range(1, sheet.max_row + 2):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value:
            normalized_value = re.sub(r'\s+', '', str(cell_value))
            if COMPANY_LABEL in normalized_value:
                start_row = row
            elif REMARKS_LABEL in normalized_value and start_row != -1:
                sections.append((start_row, row))
                start_row = -1

    if not sections:
        return None, f"'{sheet_name}' 시트에서 '{COMPANY_LABEL}' 또는 '{REMARKS_LABEL}' 라벨을 찾을 수 없어 데이터 구조를 인식할 수 없습니다."

    # 2. 삽입 위치 탐색
    target_row, target_col = -1, -1
    found_empty_slot = False
    for sec_start, sec_end in sections:
        company_row = next((r for r in range(sec_start, sec_end + 1) if COMPANY_LABEL in re.sub(r'\s+', '', str(sheet.cell(row=r, column=1).value))), -1)
        if company_row == -1: continue

        for col in range(START_COL, END_COL + 1):
            cell = sheet.cell(row=company_row, column=col)
            if cell.value is None or str(cell.value).strip() in ["", "-", "—", "N/A", "없음"]:
                target_row = sec_start
                target_col = col
                found_empty_slot = True
                break
        if found_empty_slot: break

    # 3. 새 구간 생성
    if not found_empty_slot:
        last_section_start, last_section_end = sections[-1]
        section_height = last_section_end - last_section_start
        target_row = last_section_end + 1
        target_col = START_COL
        NO_FILL = PatternFill(fill_type=None) # 채우기 없음 서식 정의

        # 행 높이 복사 및 전체 열 서식 복사
        for i in range(section_height + 1):
            source_row, new_row = last_section_start + i, target_row + i
            
            # 행 높이 복사
            if sheet.row_dimensions[source_row].height:
                sheet.row_dimensions[new_row].height = sheet.row_dimensions[source_row].height

            # A열부터 M열까지 서식 복사
            for col in range(1, END_COL + 1):
                source_cell = sheet.cell(row=source_row, column=col)
                new_cell = sheet.cell(row=new_row, column=col)

                # 서식 복사
                if source_cell.has_style:
                    new_cell.font = source_cell.font.copy()
                    new_cell.border = source_cell.border.copy()
                    new_cell.number_format = source_cell.number_format
                    new_cell.protection = source_cell.protection.copy()
                    new_cell.alignment = source_cell.alignment.copy()
                    
                    # [핵심 수정] A열만 채우기(fill)를 복사하고, B~M열은 채우기 없음으로 설정
                    if col == 1:
                        new_cell.fill = source_cell.fill.copy()
                    else:
                        new_cell.fill = NO_FILL

                # A열(라벨)은 값도 복사, 나머지는 값 비우기
                if col == 1:
                    new_cell.value = source_cell.value
                else:
                    new_cell.value = None # 서식만 복사하고 값은 비움
        
        # 새 구간을 sections 리스트에 추가하여 아래 데이터 삽입 로직이 인식하도록 함
        sections.append((target_row, target_row + section_height))

    # 4. 데이터 삽입
    labels_in_section = {}
    current_section_start_row = next((s_start for s_start, s_end in sections if s_start <= target_row <= s_end), -1)
    if current_section_start_row == -1 and target_row > sections[-1][1]:
        current_section_start_row = target_row
    
    if current_section_start_row != -1:
        sec_end = next((s_end for s_start, s_end in sections if s_start == current_section_start_row), sheet.max_row)
        for r in range(current_section_start_row, sec_end + 1):
            cell_val = sheet.cell(row=r, column=1).value
            if cell_val:
                labels_in_section[re.sub(r'\s+', '', str(cell_val))] = r

    final_data = new_data.copy()
    final_data['상호'] = company_name

    for key, value in final_data.items():
        excel_label_from_map = COLUMN_MAP.get(key)
        if excel_label_from_map:
            normalized_label_from_map = re.sub(r'\s+', '', excel_label_from_map)
            
            # '상호'는 '회사명'으로, '시공능력'은 공종에 맞게 동적으로 처리
            if normalized_label_from_map == '상호':
                final_excel_label = '회사명'
            elif normalized_label_from_map == '시공능력':
                final_excel_label = db_type + '시공능력'
            else:
                final_excel_label = normalized_label_from_map

            if final_excel_label in labels_in_section:
                row_to_insert = labels_in_section[final_excel_label]
                cell = sheet.cell(row=row_to_insert, column=target_col)
                
                # --- [핵심] 서식 적용 ---
                cell.font = DEFAULT_FONT # 기본 폰트 먼저 적용
                if final_excel_label == '회사명':
                    cell.fill = GREY_FILL
                    cell.font = BOLD_FONT
                else:
                    cell.fill = GREEN_FILL

                if value:
                    value_str = str(value).replace(",", "").replace("%", "")
                    try:
                        if '비율' in key:
                            cell.value = float(value_str) / 100.0
                            cell.number_format = '0.00%'
                        elif key in ['시평액', '3년실적', '5년실적']:
                            cell.value = int(float(value_str)) * 1000
                        else:
                            cell.value = value
                    except (ValueError, TypeError):
                        cell.value = value
    try:
        workbook.save(excel_path)
        return f"'{sheet_name}' 시트에 신규 업체 '{company_name}' 정보를 추가했습니다.", None
    except Exception as e:
        return None, f"엑셀 파일 저장 오류: {e}"


def batch_update_colors(excel_path):
    try:
        workbook = load_workbook(filename=excel_path)
    except Exception as e:
        return f"엑셀 파일 열기 오류: {e}"
    GREEN_COLOR = Color(type='theme', theme=6, tint=0.7999816888943144)
    BLUE_COLOR = Color(type='theme', theme=3, tint=0.7999816888943144)
    NEW_BLUE_FILL = PatternFill(fgColor=BLUE_COLOR, fill_type="solid")
    NO_FILL = PatternFill(fill_type=None)
    update_count = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=2):
            label_cell = row[0]
            if label_cell.value and '신용평가' in str(label_cell.value): continue
            for cell in row[1:]:
                current_color = cell.fill.fgColor if cell.fill else None
                if cell.value is None or str(cell.value).strip() == "":
                    if cell.fill and cell.fill.fill_type is not None:
                        cell.fill = NO_FILL; update_count += 1
                else:
                    if current_color == GREEN_COLOR:
                        cell.fill = NEW_BLUE_FILL; update_count += 1
                    elif current_color == BLUE_COLOR:
                        cell.fill = NO_FILL; update_count += 1
    try:
        workbook.save(excel_path)
        return f"총 {update_count}개 셀의 서식을 성공적으로 업데이트했습니다."
    except Exception as e:
        return f"엑셀 파일 저장 오류: {e}"


def update_credit_rating_only(excel_path, biz_no_to_find, new_credit_rating):
    try:
        workbook = load_workbook(filename=excel_path)
    except Exception as e:
        return None, f"엑셀 파일 열기 오류: {e}"
    THEME_GREEN_COLOR = Color(type='theme', theme=6, tint=0.7999816888943144)
    GREEN_FILL = PatternFill(fgColor=THEME_GREEN_COLOR, fill_type="solid")
    target_row, target_col, target_sheet_name = -1, -1, None; found = False
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None: continue
                if str(cell.value).strip().replace('-', '') == biz_no_to_find.strip().replace('-', ''):
                    target_row, target_col, target_sheet_name = cell.row, cell.column, sheet_name
                    found = True; break
            if found: break
        if found: break
    if not found: return f"해당 업체를 찾을 수 없습니다.", None
    sheet = workbook[target_sheet_name]
    credit_rating_offset = RELATIVE_OFFSETS.get('신용평가')
    if credit_rating_offset is None: return None, "'config.py'의 RELATIVE_OFFSETS에 '신용평가'가 정의되지 않았습니다."
    update_row = target_row + credit_rating_offset
    if not (1 <= update_row <= sheet.max_row and 1 <= target_col <= sheet.max_column): return None, f"'신용평가' 셀의 위치({update_row}행)가 유효하지 않습니다."
    cell_to_update = sheet.cell(row=update_row, column=target_col)
    cell_to_update.value = new_credit_rating
    cell_to_update.fill = GREEN_FILL
    try:
        workbook.save(excel_path)
        return "업데이트 완료!", None
    except Exception as e:
        return None, f"엑셀 파일 저장 오류: {e}"


def batch_update_credit_rating_colors(excel_path):
    try:
        workbook = load_workbook(filename=excel_path)
    except Exception as e:
        return f"엑셀 파일 열기 오류: {e}"
    GREEN_FILL = PatternFill(fgColor=Color(type='theme', theme=6, tint=0.7999816888943144), fill_type="solid")
    BLUE_FILL = PatternFill(fgColor=Color(type='theme', theme=3, tint=0.7999816888943144), fill_type="solid")
    NO_FILL = PatternFill(fill_type=None)
    today = datetime.now().date()
    update_count = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=2):
            label_cell = row[0]
            if label_cell.value and '신용평가' in str(label_cell.value):
                for cell in row[1:]:
                    if cell.value is None or str(cell.value).strip() == "":
                        cell.fill = NO_FILL; update_count += 1; continue
                    match = re.search(r'~(\d{2,4}\.\d{2}\.\d{2})', str(cell.value))
                    if not match: continue
                    end_date_str = match.group(1)
                    try:
                        expiry_date = datetime.strptime(end_date_str, '%y.%m.%d').date()
                        if expiry_date < today: cell.fill = BLUE_FILL
                        else: cell.fill = GREEN_FILL
                        update_count += 1
                    except ValueError:
                        continue
    try:
        workbook.save(excel_path)
        return f"총 {update_count}개의 신용평가 셀 색상을 갱신했습니다."
    except Exception as e:
        return f"엑셀 파일 저장 오류: {e}"
